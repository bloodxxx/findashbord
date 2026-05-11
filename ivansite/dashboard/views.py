import json
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
import csv

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os

from .models import Document, FinancialRecord, AnalysisResult, Metric
from .parsers import parse_document
from .analyzers import analyze_document


ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv', 'xml'}
MAX_FILE_SIZE = 20 * 1024 * 1024 


def login_view(request):
    # авторизация пользователя
    if request.user.is_authenticated:
        return redirect('dashboard:upload')
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard:upload')
        messages.error(request, 'Неверный логин или пароль')
    return render(request, 'dashboard/login.html')


def register_view(request):
    # регистрация нового пользователя
    if request.user.is_authenticated:
        return redirect('dashboard:upload')
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        password2 = request.POST.get('password2', '')
        if not username or not password:
            messages.error(request, 'Заполните все поля')
        elif password != password2:
            messages.error(request, 'Пароли не совпадают')
        elif len(password) < 6:
            messages.error(request, 'Пароль слишком короткий (минимум 6 символов)')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'Пользователь с таким именем уже существует')
        else:
            User.objects.create_user(username=username, password=password)
            messages.success(request, 'Аккаунт создан. Войдите в систему.')
            return redirect('dashboard:login')
    return render(request, 'dashboard/register.html')


def logout_view(request):
    # выход из системы
    logout(request)
    return redirect('dashboard:login')


@login_required
def upload_view(request):
    # загрузка и обработка документа
    if request.method == 'POST':
        file = request.FILES.get('file')
        doc_type = request.POST.get('doc_type', '')
        period = request.POST.get('period', '')

        if not file:
            messages.error(request, 'Выберите файл')
            return redirect('dashboard:upload')

        ext = file.name.rsplit('.', 1)[-1].lower() if '.' in file.name else ''
        if ext not in ALLOWED_EXTENSIONS:
            messages.error(request, f'Неподдерживаемый формат. Допустимые: {", ".join(ALLOWED_EXTENSIONS)}')
            return redirect('dashboard:upload')

        if file.size > MAX_FILE_SIZE:
            messages.error(request, 'Файл слишком большой (максимум 20 МБ)')
            return redirect('dashboard:upload')

        if doc_type not in dict(Document.DOCUMENT_TYPES):
            messages.error(request, 'Выберите тип документа')
            return redirect('dashboard:upload')

        document = Document.objects.create(
            user=request.user,
            file=file,
            file_name=file.name,
            doc_type=doc_type,
            period=period,
            status=Document.STATUS_VALIDATING,
        )

        try:
            document.file.seek(0)
            records_data = parse_document(document.file, document.file_name, doc_type)

            document.status = Document.STATUS_ANALYZING
            document.save()

            FinancialRecord.objects.filter(document=document).delete()
            for rd in records_data:
                FinancialRecord.objects.create(
                    document=document,
                    date=rd.get('date'),
                    category=rd.get('category', ''),
                    amount=rd.get('amount', 0),
                    record_type=rd.get('record_type', ''),
                    counterparty=rd.get('counterparty', ''),
                    department=rd.get('department', ''),
                    indicator=rd.get('indicator', ''),
                    plan_value=rd.get('plan_value'),
                    fact_value=rd.get('fact_value'),
                    period=rd.get('period', ''),
                )

            analyze_document(document)
            document.status = Document.STATUS_DONE
            document.save()
            messages.success(request, f'Документ «{document.file_name}» успешно проанализирован')
            return redirect('dashboard:document_detail', pk=document.pk)

        except ValueError as e:
            document.status = Document.STATUS_ERROR
            document.error_message = str(e)
            document.save()
            messages.error(request, f'Ошибка: {e}')
            return redirect('dashboard:upload')
        except Exception as e:
            document.status = Document.STATUS_ERROR
            document.error_message = str(e)
            document.save()
            messages.error(request, 'Произошла ошибка при обработке файла')
            return redirect('dashboard:upload')

    recent_docs = Document.objects.filter(user=request.user)[:5]
    return render(request, 'dashboard/upload.html', {
        'document_types': Document.DOCUMENT_TYPES,
        'recent_docs': recent_docs,
    })


@login_required
def history_view(request):
    # история загруженных документов с фильтрацией и пагинацией
    docs_qs = Document.objects.filter(user=request.user)

    doc_type_filter = request.GET.get('doc_type', '')
    status_filter = request.GET.get('status', '')
    period_filter = request.GET.get('period', '')

    if doc_type_filter:
        docs_qs = docs_qs.filter(doc_type=doc_type_filter)
    if status_filter:
        docs_qs = docs_qs.filter(status=status_filter)
    if period_filter:
        docs_qs = docs_qs.filter(period__icontains=period_filter)

    paginator = Paginator(docs_qs, 10)
    page = request.GET.get('page', 1)
    documents = paginator.get_page(page)

    return render(request, 'dashboard/history.html', {
        'documents': documents,
        'document_types': Document.DOCUMENT_TYPES,
        'statuses': Document.STATUSES,
        'current_filters': {
            'doc_type': doc_type_filter,
            'status': status_filter,
            'period': period_filter,
        },
    })


@login_required
def document_detail(request, pk):
    # детальный просмотр документа и результатов анализа
    document = get_object_or_404(Document, pk=pk, user=request.user)
    analysis = getattr(document, 'analysis', None)
    metrics = document.metrics.all()
    records = document.records.all()[:100]

    chart_data_json = json.dumps(analysis.chart_data if analysis else {})

    return render(request, 'dashboard/document_detail.html', {
        'document': document,
        'analysis': analysis,
        'metrics': metrics,
        'records': records,
        'chart_data_json': chart_data_json,
    })


@login_required
def delete_document(request, pk):
    # удаление документа пользователя
    document = get_object_or_404(Document, pk=pk, user=request.user)
    if request.method == 'POST':
        document.delete()
        messages.success(request, 'Документ удалён')
    return redirect('dashboard:history')


@login_required
def export_csv(request, pk):
    # экспорт записей документа в CSV файл
    document = get_object_or_404(Document, pk=pk, user=request.user)
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="report_{document.pk}.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)

    if document.doc_type == 'income_expense':
        writer.writerow(['Дата', 'Статья', 'Тип', 'Сумма'])
        for r in document.records.all():
            writer.writerow([r.date, r.category, r.record_type, r.amount])
    elif document.doc_type == 'cash_flow':
        writer.writerow(['Дата', 'Тип', 'Контрагент', 'Сумма'])
        for r in document.records.all():
            writer.writerow([r.date, r.record_type, r.counterparty, r.amount])
    elif document.doc_type == 'budget':
        writer.writerow(['Период', 'Подразделение', 'Статья', 'План', 'Факт', 'Отклонение'])
        for r in document.records.all():
            dev = (r.fact_value or 0) - (r.plan_value or 0)
            writer.writerow([r.period, r.department, r.category, r.plan_value, r.fact_value, dev])
    elif document.doc_type == 'kpi':
        writer.writerow(['Период', 'Подразделение', 'Показатель', 'План', 'Факт', 'Отклонение'])
        for r in document.records.all():
            dev = (r.fact_value or 0) - (r.plan_value or 0)
            writer.writerow([r.period, r.department, r.indicator, r.plan_value, r.fact_value, dev])

    return response


@login_required
def export_excel(request, pk):
    # экспорт записей документа в Excel (.xlsx)
    document = get_object_or_404(Document, pk=pk, user=request.user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    # стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", indent=1)
    center = Alignment(horizontal="center", vertical="center")
    right  = Alignment(horizontal="right",  vertical="center", indent=1)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row_num, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

    def style_data_cell(cell, align=None):
        cell.border = border
        cell.alignment = align or center

    # заголовок документа
    ws.merge_cells(f"A1:{chr(64 + 6)}1")
    title_cell = ws["A1"]
    title_cell.value = f"Отчёт: {document.file_name}"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{chr(64 + 6)}2")
    sub_cell = ws["A2"]
    sub_cell.value = f"{document.get_doc_type_display()} · {document.period or ''} · {document.uploaded_at.strftime('%d.%m.%Y')}"
    sub_cell.font = Font(italic=True, color="6B7280", size=10)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    data_start = 4  # строка с заголовками таблицы

    if document.doc_type == 'income_expense':
        headers = ['Дата', 'Статья', 'Тип', 'Сумма']
        ws.append([])
        for i, h in enumerate(headers, 1):
            ws.cell(row=data_start, column=i).value = h
        style_header_row(ws, data_start, len(headers))
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 16
        for row_idx, r in enumerate(document.records.all(), data_start + 1):
            ws.cell(row=row_idx, column=1, value=str(r.date) if r.date else '').alignment = center
            ws.cell(row=row_idx, column=2, value=r.category or '').alignment = left
            ws.cell(row=row_idx, column=3, value='Доход' if r.record_type == 'income' else 'Расход').alignment = center
            amt = ws.cell(row=row_idx, column=4, value=float(r.amount) if r.amount else 0)
            amt.number_format = '#,##0.00'
            amt.alignment = right
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).border = border

    elif document.doc_type == 'cash_flow':
        headers = ['Дата', 'Тип', 'Контрагент', 'Сумма']
        ws.append([])
        for i, h in enumerate(headers, 1):
            ws.cell(row=data_start, column=i).value = h
        style_header_row(ws, data_start, len(headers))
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 16
        for row_idx, r in enumerate(document.records.all(), data_start + 1):
            ws.cell(row=row_idx, column=1, value=str(r.date) if r.date else '').alignment = center
            ws.cell(row=row_idx, column=2, value='Поступление' if r.record_type == 'inflow' else 'Списание').alignment = center
            ws.cell(row=row_idx, column=3, value=r.counterparty or '').alignment = left
            amt = ws.cell(row=row_idx, column=4, value=float(r.amount) if r.amount else 0)
            amt.number_format = '#,##0.00'
            amt.alignment = right
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).border = border

    elif document.doc_type == 'budget':
        headers = ['Период', 'Подразделение', 'Статья', 'План', 'Факт', 'Отклонение']
        ws.append([])
        for i, h in enumerate(headers, 1):
            ws.cell(row=data_start, column=i).value = h
        style_header_row(ws, data_start, len(headers))
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 28
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        for row_idx, r in enumerate(document.records.all(), data_start + 1):
            dev = (r.fact_value or 0) - (r.plan_value or 0)
            ws.cell(row=row_idx, column=1, value=r.period or '').alignment = center
            ws.cell(row=row_idx, column=2, value=r.department or '').alignment = left
            ws.cell(row=row_idx, column=3, value=r.category or '').alignment = left
            for col, val in [(4, r.plan_value), (5, r.fact_value), (6, dev)]:
                c = ws.cell(row=row_idx, column=col, value=float(val) if val is not None else 0)
                c.number_format = '#,##0.00'
                c.alignment = right
                if col == 6:
                    c.font = Font(color="059669" if dev >= 0 else "DC2626", bold=True)
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).border = border

    elif document.doc_type == 'kpi':
        headers = ['Период', 'Подразделение', 'Показатель', 'План', 'Факт', 'Отклонение']
        ws.append([])
        for i, h in enumerate(headers, 1):
            ws.cell(row=data_start, column=i).value = h
        style_header_row(ws, data_start, len(headers))
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 28
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        for row_idx, r in enumerate(document.records.all(), data_start + 1):
            dev = (r.fact_value or 0) - (r.plan_value or 0)
            ws.cell(row=row_idx, column=1, value=r.period or '').alignment = center
            ws.cell(row=row_idx, column=2, value=r.department or '').alignment = left
            ws.cell(row=row_idx, column=3, value=r.indicator or '').alignment = left
            for col, val in [(4, r.plan_value), (5, r.fact_value), (6, dev)]:
                c = ws.cell(row=row_idx, column=col, value=float(val) if val is not None else 0)
                c.number_format = '#,##0.00'
                c.alignment = right
                if col == 6:
                    c.font = Font(color="059669" if dev >= 0 else "DC2626", bold=True)
            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).border = border

    # фиксируем строку заголовков
    ws.freeze_panes = f"A{data_start + 1}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="report_{document.pk}.xlsx"'
    return response


@login_required
def export_pdf(request, pk):
    # экспорт записей документа в PDF
    document = get_object_or_404(Document, pk=pk, user=request.user)

    # регистрируем шрифт с поддержкой кириллицы
    font_path = os.path.join(os.path.dirname(__file__), 'fonts', 'DejaVuSans.ttf')
    font_bold_path = os.path.join(os.path.dirname(__file__), 'fonts', 'DejaVuSans-Bold.ttf')
    pdfmetrics.registerFont(TTFont('DejaVu', font_path))
    pdfmetrics.registerFont(TTFont('DejaVu-Bold', font_bold_path))

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', fontName='DejaVu-Bold', fontSize=14, alignment=1, spaceAfter=4)
    sub_style = ParagraphStyle('sub', fontName='DejaVu', fontSize=9, alignment=1, textColor=colors.HexColor('#6B7280'), spaceAfter=14)
    cell_style = ParagraphStyle('cell', fontName='DejaVu', fontSize=8)

    elements = []

    elements.append(Paragraph(f"Отчёт: {document.file_name}", title_style))
    elements.append(Paragraph(
        f"{document.get_doc_type_display()} · {document.period or ''} · {document.uploaded_at.strftime('%d.%m.%Y')}",
        sub_style
    ))

    HEADER_COLOR = colors.HexColor('#4F46E5')
    ROW_ALT_COLOR = colors.HexColor('#F5F3FF')
    POS_COLOR = colors.HexColor('#059669')
    NEG_COLOR = colors.HexColor('#DC2626')

    def make_table(headers, rows_data, text_cols=None, num_cols=None, col_widths=None):
        """
        text_cols  — индексы колонок с текстом (0-based), выравнивание LEFT
        num_cols   — индексы числовых колонок, выравнивание RIGHT
        col_widths — список долей ширины для каждой колонки (сумма = 1.0),
                     если None — равные доли
        """
        text_cols = set(text_cols or [])
        num_cols  = set(num_cols  or [])

        page_width = landscape(A4)[0] - 3 * cm

        if col_widths:
            widths = [page_width * w for w in col_widths]
        else:
            widths = [page_width / len(headers)] * len(headers)

        # стили параграфов для ячеек данных
        p_left   = ParagraphStyle('pl', fontName='DejaVu', fontSize=8,
                                  alignment=0, leading=10, wordWrap='LTR')
        p_center = ParagraphStyle('pc', fontName='DejaVu', fontSize=8,
                                  alignment=1, leading=10, wordWrap='LTR')
        p_right  = ParagraphStyle('pr', fontName='DejaVu', fontSize=8,
                                  alignment=2, leading=10, wordWrap='LTR')
        h_style  = ParagraphStyle('h',  fontName='DejaVu-Bold', fontSize=8,
                                  textColor=colors.white, alignment=1, leading=10)

        def cell_style(col_idx):
            if col_idx in text_cols:
                return p_left
            if col_idx in num_cols:
                return p_right
            return p_center

        # заголовки
        table_data = [[Paragraph(h, h_style) for h in headers]]
        # данные — оборачиваем в Paragraph для переноса
        for row in rows_data:
            table_data.append([
                Paragraph(str(cell) if cell is not None else '—', cell_style(ci))
                for ci, cell in enumerate(row)
            ])

        t = Table(table_data, colWidths=widths, repeatRows=1)
        style_cmds = [
            ('BACKGROUND',    (0, 0), (-1,  0), HEADER_COLOR),
            ('FONTNAME',      (0, 0), (-1,  0), 'DejaVu-Bold'),
            ('FONTNAME',      (0, 1), (-1, -1), 'DejaVu'),
            ('FONTSIZE',      (0, 0), (-1, -1), 8),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#D1D5DB')),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, ROW_ALT_COLOR]),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 5),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 5),
        ]
        t.setStyle(TableStyle(style_cmds))
        return t

    def fmt(val):
        if val is None:
            return '—'
        try:
            return f"{float(val):,.2f}"
        except Exception:
            return str(val)

    if document.doc_type == 'income_expense':
        # Дата(0)-center  Статья(1)-left  Тип(2)-center  Сумма(3)-right
        rows = []
        for r in document.records.all():
            rows.append([
                str(r.date) if r.date else '—',
                r.category or '—',
                'Доход' if r.record_type == 'income' else 'Расход',
                fmt(r.amount),
            ])
        elements.append(make_table(
            ['Дата', 'Статья', 'Тип', 'Сумма'], rows,
            text_cols=[1], num_cols=[3],
            col_widths=[0.13, 0.52, 0.13, 0.22]
        ))

    elif document.doc_type == 'cash_flow':
        # Дата(0)-center  Тип(1)-center  Контрагент(2)-left  Сумма(3)-right
        rows = []
        for r in document.records.all():
            rows.append([
                str(r.date) if r.date else '—',
                'Поступление' if r.record_type == 'inflow' else 'Списание',
                r.counterparty or '—',
                fmt(r.amount),
            ])
        elements.append(make_table(
            ['Дата', 'Тип', 'Контрагент', 'Сумма'], rows,
            text_cols=[2], num_cols=[3],
            col_widths=[0.13, 0.15, 0.50, 0.22]
        ))

    elif document.doc_type == 'budget':
        # Период(0)-center  Подразделение(1)-left  Статья(2)-left  План(3)-right  Факт(4)-right  Откл(5)-right
        rows = []
        for r in document.records.all():
            dev = (r.fact_value or 0) - (r.plan_value or 0)
            rows.append([
                r.period or '—',
                r.department or '—',
                r.category or '—',
                fmt(r.plan_value),
                fmt(r.fact_value),
                fmt(dev),
            ])
        elements.append(make_table(
            ['Период', 'Подразделение', 'Статья', 'План', 'Факт', 'Отклонение'], rows,
            text_cols=[1, 2], num_cols=[3, 4, 5],
            col_widths=[0.10, 0.20, 0.38, 0.11, 0.11, 0.10]
        ))

    elif document.doc_type == 'kpi':
        # Период(0)-center  Подразделение(1)-left  Показатель(2)-left  План(3)-right  Факт(4)-right  Откл(5)-right
        rows = []
        for r in document.records.all():
            dev = (r.fact_value or 0) - (r.plan_value or 0)
            rows.append([
                r.period or '—',
                r.department or '—',
                r.indicator or '—',
                fmt(r.plan_value),
                fmt(r.fact_value),
                fmt(dev),
            ])
        elements.append(make_table(
            ['Период', 'Подразделение', 'Показатель', 'План', 'Факт', 'Отклонение'], rows,
            text_cols=[1, 2], num_cols=[3, 4, 5],
            col_widths=[0.10, 0.20, 0.38, 0.11, 0.11, 0.10]
        ))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="report_{document.pk}.pdf"'
    return response


@login_required
def formats_view(request):
    # страница с описанием поддерживаемых форматов
    return render(request, 'dashboard/formats.html')
